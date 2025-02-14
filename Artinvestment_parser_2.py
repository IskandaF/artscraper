#!/usr/bin/env python
# coding: utf-8

from urllib import request
from PIL import Image
import PIL.ImageOps  
import time


import requests


import re


from bs4 import BeautifulSoup as bs


from openpyxl import Workbook


from openpyxl import load_workbook


import os


from skimage.io import imread


import re


datasetpath="/Users/soulist/Projects/dataset.xlsx"
imagedatasetpath="Users/soulist/Projects/imagedataset" 
wb=load_workbook("/Users/soulist/Projects/dataset.xlsx")
ws=wb.active
url="https://artinvestment.ru/auctions/"
row=1



lastpage=482
def search (url=url):
    artistid=0
    paintingid=0
    page=1
    rmain=requests.get(url)
#                        +str(page))
#                        +"&citems=5")
    soupmain=bs(rmain.text)

    while page<=lastpage:
        artistlist=soupmain.findAll("div",{"class":"artists-list"})
        painterlinks=[]
# Listing through the artists on the page (Проходимся по всем художникам на странице)
        for i in artistlist:   
            artistid+=1
            print(artistid)
            print(i.a.text,i.a["href"])
            artistname=i.a.text
            print(paintings(i.a["href"]+"/works.html",artistid,artistname)) 
#         Switching the page when no artists left
        page+=1
        rmain=requests.get("https://artinvestment.ru/auctions/?page="+str(page))
#                            +"&citems=5")
        soupmain=bs(rmain.text)   
        print("Страница - "+page)
    return ("Следующий")
def paintings (link,artistid,artistname):
        start = time.clock()
        listitem="list-item"
        page=1
        paintingid=0
        paintingartisauctiontid=0
        
        r=requests.get(link)
        soup=bs(r.text)
        print(link)
#   Looking for class "list-item" on the page. If present, then some painting is present.
# Ищем класс "list-item" на страничке. Если он присутствует, то какая-либо картина также присутствует.
        while listitem in r.text:

            global row
#             Looking for the painting's title (Ищем название картины)
            try:
                listitems=soup.find("div",{"class":"list-items"})
            except AttributeError:
                return
            try:
                paintingslist=listitems.findAll("div",{"class":"list-item"})
            except AttributeError:
                return
            for i in paintingslist:  
                    auctionid=1
                    column=7
                    paintingid+=1
                    print(paintingid)
                    paintingartisauctiontid=str(artistid)+str(paintingid)+str(auctionid)
                    ws["A"+str(row)]=paintingartisauctiontid
                    ws["B"+str(row)]=artistname
                    
                    if i.h3.text=="Перейти к работе":
                        print ("Безымянная")
                        ws["C"+str(row)]="Безымянная"
                    else:
                        print (i.h3.text)
                        ws["C"+str(row)]=i.h3.text
                    
                    print("Следующая картина")
#             Assigning year,size and technique characteristics to variables (Назначаем переменные с данными о годе создания, размере и технике)
                    year=re.findall(r"\d{4}",i.text)
                    size=re.findall(r".+\d+х\d+.+",i.text)
                    technique=re.findall(r".+;.+",i.text)

#             Printing out and writing the characteristics in excel (Распечатываем характеристики и записываем их в Эксель)
                    if len(size)!=0:
                        sizestr="".join(size)
                        print ("Размер -",sizestr)
                        sizestr.replace("Размер, ","")
                        ws["E"+str(row)]="".join(size)
                    if len(technique)!=0:
                        print ("Техника -","".join(technique))
                        ws["F"+str(row)]="".join(technique)
                    header=i.find("div",{"class":"header"})
                    print(paintingartisauctiontid)
# Finding the year of creation (Находим год создания и записываем в Excel)
                    values=i.findAll("div",{"class":"value"})
                    for v in values:
                        try:
                            year="".join(re.findall(r"^\d{4}",v.get_text()))
                            if len(year)>0:
                                print (year)
                                ws["D"+str(row)]="".join(year)
                        except AttributeError:
                            continue
#             Risk Labeling (Ищем маркировку риска и записываем в Excel)
                    labelyellow=i.find("span",{"class":"yellow round"})
                    labelgreen=i.find("span",{"class":"green round"})
                    labelred=i.find("span",{"class":"red round"})
                    label=[]
                    if labelyellow is not None:
                        label=[]
                        label.append(labelyellow["class"])
                    if labelgreen is not None:
                        label=[]
                        label.append(labelgreen["class"])
                    if labelred is not None: 
                        label=[]
                        label.append(labelred["class"])
                    try:

                        ws["L"+str(row)]=("".join(label[0]))
                    except IndexError:
                        ws["L"+str(row)]="No Label"

    #                 Downloading an image (Скачиваем изображение)
                    linktoimage=i.find("img")
                    imagepath=("/Users/soulist/Projects/imagedataset/"+str(paintingartisauctiontid)+".jpg")
                    try:
                        imageurl=linktoimage["src"].replace("small","big")    
                        request.urlretrieve(imageurl,imagepath)
                        ws["M"+str(row)]=imageurl
                    except KeyboardInterrupt:
                        raise
                    except:
                        continue
    #         Inverting the negative (Инвертируем цвета)
    #                 image = Image.open(imagepath)
    #                 inverted_image = PIL.ImageOps.invert(image)
    #                 inverted_image.save(imagepath)
    
    
# Collecting the data about last auction (Собираем данные о последнем аукционе)
                    for elm in header.find_next_sibling("div").find_all("div"):
                        print(elm.getText())
                        ws.cell(row=row,column=column).value=elm.get_text()
                        column+=1
                    column=7
                    row+=1
                    wb.save("/Users/soulist/Projects/dataset.xlsx")
#   Looking for the second auction information (Ищем информацию о втором аукционе) 
                    try:
                        print ("Следующий аукцион")
                        for elm in header.find_next_sibling("div").find_next_sibling("div").find_all("div"):
                            print(elm.get_text())
                            ws.cell(row=row,column=column).value=elm.get_text()
                            column+=1
#       If second auction is present, writing down all previous metadata on the new row (
#  Если информация о втором аукционе есть, записываем все метаданные по новой на вторую строчку вместе с данными второго аукциона
                        if i.h3.text=="Перейти к работе":
                            print ("Безымянная")
                            ws["B"+str(row)]="Безымянная"
                        else:
                            print (i.h3.text)
                            ws["B"+str(row)]=i.h3.text
                        for v in values:
    #                     print(v.getText())
    #                     print("".join(re.findall(r"^\d{4}$",v.get_text())))
                            year="".join(re.findall(r"^\d{4}",v.get_text()))
                            if len(year)>0:
                                print (year)
                                ws["C"+str(row)]="".join(year)
                        if len(size)!=0:
                            sizestr="".join(size)
                            sizestr.replace("Размер, ","")
                            ws["D"+str(row)]="".join(size)
                        if len(technique)!=0:
                            ws["E"+str(row)]="".join(technique)
    #                     Risk Labeling
                        label=[]
                        labelyellow=i.find("span",{"class":"yellow round"})
                        labelgreen=i.find("span",{"class":"green round"})
                        labelred=i.find("span",{"class":"red round"})

                        if labelyellow is not None:
                            label=[]
                            label.append(labelyellow["class"])
                        if labelgreen is not None:
                            label=[]
                            label.append(labelgreen["class"])
                        if labelred is not None: 
                            label=[]
                            label.append(labelred["class"])
            #             label1=0
            #             label1=list(filter(None, label))
                        try:
                            ws["K"+str(row)]=("".join(label[0]))
                        except IndexError:
                            ws["K"+str(row)]="No Label"

                        try:
                            imageurl=linktoimage["src"].replace("small","big")    
                            request.urlretrieve(imageurl,imagepath)
                            ws["L"+str(row)]=imageurl
                        except KeyboardInterrupt:
                            raise
                        except:
                            continue
#       Numbering all ids one number up (Прибавляем ко всем номерам 1)
                        ws["B"+str(row)]=artistname
                        auctionid+=1
                        paintingartisauctiontid=str(artistid)+str(paintingid)+str(auctionid)
                        ws["A"+str(row)]=paintingartisauctiontid
                        row+=1
#                     Saving the Excel file (Сохраняем в эксель)
                        wb.save("/Users/soulist/Projects/dataset.xlsx")
                    except AttributeError:
                        continue

                    
#           If no more paintings on the page, go to the next page 
# Если больше картин на странице нет, идём на следующую 
            page+=1
            print(page)
            r=requests.get(link+"?page="+str(page))
            soup=bs(r.text)
            print(link+"?page="+str(page))

            print(link)
        wb.save("/Users/soulist/Projects/dataset.xlsx")
        return
print(search())
row=1
# paintings("https://artinvestment.ru/auctions/2126/works.html",1,"Авилов")




